#!/usr/bin/env python3
"""Fetch BLS OEWS May 2025 national industry-specific estimates (oesm25in4.zip).

Repricing Main Street, dataset 1 of 3 (engine backbone).
Target dir: 02_data/raw/2026-06-10_oews/
Run from anywhere: python3 03_engine/scripts/fetch_oews.py [--force]

Stdlib only. If openpyxl is installed, also verifies a NAICS 5412 row inside
the 4-digit industry XLSX and counts its data rows.

On success, appends a Landing record to the folder's PROVENANCE.md.
Written by DATA-2 agent, 2026-06-10. Sandbox egress to bls.gov was blocked,
so this script is the official landing path (run on William's Mac).
"""

import hashlib
import re
import sys
import urllib.request
import zipfile
from datetime import datetime, timezone
from pathlib import Path

URL = "https://www.bls.gov/oes/special-requests/oesm25in4.zip"
PROJECT_ROOT = Path(__file__).resolve().parents[2]
OUT_DIR = PROJECT_ROOT / "02_data" / "raw" / "2026-06-10_oews"
ZIP_PATH = OUT_DIR / "oesm25in4.zip"
EXTRACT_DIR = OUT_DIR / "oesm25in4"
PROVENANCE = OUT_DIR / "PROVENANCE.md"
# BLS blocks default urllib user agents; identify as a browser plus contact.
USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0 Safari/537.36 "
    "RepricingMainStreet-research/1.0 (contact: sheets@egatewaycapital.com)"
)


def utc_now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def download(url: str, dest: Path) -> int:
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    done = 0
    try:
        with urllib.request.urlopen(req, timeout=180) as resp, open(dest, "wb") as out:
            status = getattr(resp, "status", 200)
            if status != 200:
                raise RuntimeError(f"HTTP {status} for {url}")
            while True:
                chunk = resp.read(1 << 20)
                if not chunk:
                    break
                out.write(chunk)
                done += len(chunk)
                if done % (5 << 20) < (1 << 20):
                    print(f"  ... {done / 1e6:.1f} MB")
    except Exception as e:
        if "CERTIFICATE_VERIFY_FAILED" in str(e):
            print("TLS certs missing. Run: /Applications/Python*/Install Certificates.command")
        raise
    return done


def safe_extract(zf: zipfile.ZipFile, target: Path) -> None:
    for m in zf.namelist():
        p = (target / m).resolve()
        if not str(p).startswith(str(target.resolve())):
            raise RuntimeError(f"Unsafe zip member path: {m}")
    zf.extractall(target)


def verify_xlsx_5412(xlsx_path: Path):
    """Return (n_data_rows, example_row_dict) for NAICS 5412, or (None, None)."""
    try:
        import openpyxl  # type: ignore
    except ImportError:
        print("openpyxl not installed; skipping in-file verification.")
        print("Optional: pip3 install openpyxl, then rerun for the 5412 row check.")
        return None, None
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [str(c) if c is not None else "" for c in next(rows)]
    idx = {name.upper(): i for i, name in enumerate(header)}
    naics_i = idx.get("NAICS")
    occ_i = idx.get("OCC_CODE")
    n = 0
    example = None
    have_total_row = False
    for row in rows:
        n += 1
        if naics_i is None or have_total_row:
            continue
        naics = str(row[naics_i]).strip() if row[naics_i] is not None else ""
        if naics in ("5412", "541200"):
            d = dict(zip(header, [str(c) for c in row]))
            if example is None:
                example = d
            if occ_i is not None and str(row[occ_i]).strip() == "00-0000":
                example = d  # prefer the all-occupations total row
                have_total_row = True
    wb.close()
    return n, example


def append_provenance(text: str) -> None:
    with open(PROVENANCE, "a", encoding="utf-8") as f:
        f.write(text)


def main() -> int:
    force = "--force" in sys.argv
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    if ZIP_PATH.exists() and not force:
        print(f"REFUSING to overwrite existing {ZIP_PATH} (raw/ is immutable). Use --force only if the prior pull was bad.")
        return 1

    t0 = utc_now()
    print(f"[{t0}] downloading {URL}")
    nbytes = download(URL, ZIP_PATH)
    digest = sha256_file(ZIP_PATH)
    print(f"  saved {ZIP_PATH.name}: {nbytes:,} bytes  sha256={digest}")

    with zipfile.ZipFile(ZIP_PATH) as zf:
        members = zf.namelist()
        print(f"  zip members ({len(members)}):")
        for m in members:
            print(f"    {m}")
        four_digit = [m for m in members if re.search(r"4d.*\.xlsx$", m, re.I)]
        if not four_digit:
            print("VERIFICATION FAILED: no 4-digit industry XLSX found in zip.")
            return 2
        print(f"  4-digit industry file(s): {four_digit}")
        safe_extract(zf, EXTRACT_DIR)
        print(f"  extracted to {EXTRACT_DIR}")

    xlsx = EXTRACT_DIR / four_digit[0]
    n_rows, example = verify_xlsx_5412(xlsx)
    if n_rows is not None:
        print(f"  {four_digit[0]}: {n_rows:,} data rows")
    if example:
        print("  NAICS 5412 example row:")
        for k in list(example)[:14]:
            print(f"    {k}: {example[k]}")
    elif n_rows is not None:
        print("VERIFICATION WARNING: no NAICS 5412 row found in 4-digit file. Investigate before use.")

    rec = (
        "\n## Landing record (auto-appended by fetch_oews.py)\n\n"
        f"- Retrieved (UTC): {t0}\n"
        f"- Source URL: {URL}\n"
        f"- File: oesm25in4.zip, {nbytes:,} bytes\n"
        f"- sha256: {digest}\n"
        f"- Zip members: {len(members)} ({'; '.join(members)})\n"
        f"- 4-digit industry file: {four_digit[0]}"
        + (f", {n_rows:,} data rows\n" if n_rows is not None else " (row count pending openpyxl)\n")
        + (
            f"- NAICS 5412 verification row: NAICS={example.get('NAICS')}, OCC_CODE={example.get('OCC_CODE')}, "
            f"OCC_TITLE={example.get('OCC_TITLE')}, TOT_EMP={example.get('TOT_EMP')}\n"
            if example
            else "- NAICS 5412 verification: pending openpyxl run\n"
        )
        + "- Verified by: fetch_oews.py\n"
    )
    append_provenance(rec)
    print(f"Landing record appended to {PROVENANCE}")
    print("NEXT: flip this dataset's row in 02_data/MANIFEST.md from SCRIPT-ONLY to LANDED.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
